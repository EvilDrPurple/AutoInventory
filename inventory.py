import pathlib
import re
import time
from datetime import datetime

import config
import traceback
import PySimpleGUI as sg
from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from selenium.webdriver.chrome.service import Service as ChromeService
from splinter import Browser

URL = 'https://fedsso.yum.com/idp/startSSO.ping?PartnerSpId=https://yumph.altametrics.com/'
USER = config.USER
PASS = config.PASS
PATH = pathlib.Path(__file__).parent.resolve().__str__()
UNITS = {'EACH': {'DISK', 'EACH'},
        'BTL': 'BOTTLE',
        'GAL': 'GALLON'}
MIN_ROW = 6
MAX_ROW = 115
AUTO_SAVE = False
LEGACY = False
BROWSER_NAME = 'firefox'


def startup_gui():
    sg.theme('DarkPurple4')   
    FONT = 'Ariel 14'
    TOOLTIP = 'Uncheck this option if there is already an existing inventory sheet for the specified date and frequency in eResturant'
    OPTION_MENU = sg.OptionMenu(['Daily', 'Weekly', 'Monthly'], default_value='Weekly', key='-FREQ-')
    CALENDAR_BUTTON = sg.CalendarButton('Select Date', target='-DATE-', format='%m/%d/%Y')
    CHECKBOX = sg.Checkbox('Create new inventory sheet', font=FONT, text_color='white', default=True, key='-NEW_INV-', tooltip=TOOLTIP)

    layout = [  [sg.Text('Select count frequency:', font=FONT, text_color='white'), sg.Push(), OPTION_MENU],
                [sg.Text('Enter date (mm/dd/yyyy):', font=FONT, text_color='white'), sg.Push(), sg.Input(key='-DATE-', size=12), CALENDAR_BUTTON],
                [sg.Text('Select inventory spreadsheet:', font=FONT, text_color='white')],
                [sg.Input(key='-FILE-', size=50), sg.FileBrowse(file_types=(('Microsoft Excel Worksheet (.xlsx)', '*.xlsx'),))],
                [sg.Push(), CHECKBOX, sg.Push()],
                [sg.Text()],
                [sg.Push(), sg.Ok(font=FONT), sg.Cancel(font=FONT), sg.Push()] ]

    window = sg.Window('AutoInventory - v0.3', layout)

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Cancel':
            break

        if not vali_date(values['-DATE-']):
            sg.popup('Please enter a valid date', title='Invalid date', font=FONT, text_color='white', keep_on_top=True)
        elif not values['-FILE-'].endswith('.xlsx'):
            sg.popup('Please select a valid file', title='Invalid file', font=FONT, text_color='white', keep_on_top=True)
        else:
            window.close()
            return values['-FREQ-'], values['-DATE-'], values['-FILE-'], values['-NEW_INV-']

    window.close()


def vali_date(date):
    return re.search('^(1[0-2]|0?[1-9])/(0?[1-9]|[1-2]\d|3[0-1])/\d{4}$', date)


def wait_for_load():
    while browser.find_by_id('loading_layer').visible:
        time.sleep(0.5)


def find_and_click(items, search_type, search_text=''):
    for item in items:
        if (
            search_type == 'text' and item.text == search_text
            or search_type == 'visible' and item.visible
        ):
            item.click()
            return True
        
    return False


class Item:
    def __init__(self):
        self.item_code = None

    def __str__(self):
        return f"{self.item_code : <12}{self.item_desc : <30}{self.item_count} {self.item_unit}"

    def parse_row(self, row):
        for cell in row:
            if type(cell) is EmptyCell: continue
            match cell.column_letter:
                case 'A':
                    if not cell.value: return False
                    self.item_code = cell.value.strip()
                case 'B':
                    self.item_desc = cell.value.strip().replace('\t', '')
                    # Adjustment for Pepsi gallons
                    self.item_code = 'V62 Syrup' if self.item_desc == 'BNB PEPSI 5 GL SYRUP' else self.item_code
                    self.item_code = 'V65 Syrup' if self.item_desc == 'BNB PEPSI 3 GL SYRUP' else self.item_code
                case 'C':
                    self.item_unit = cell.value.strip()
                    # Adjustment for oregano
                    self.item_unit = 'CASE' if self.item_code == '74727' else self.item_unit
                    self.item_unit = UNITS[self.item_unit] if self.item_unit in UNITS else self.item_unit
                case 'H':
                    if not cell.value: return False
                    self.item_count = cell.value

        if not self.item_code: return False
        return True
    
    def enter_data(self):
        browser.find_by_id('INV_ACC_DETAIL_tbl_filter').find_by_tag('input').fill(self.item_code)

        for td in browser.find_by_id('INV_ACC_DETAIL_tbl').find_by_tag('td'):
            if td.text and td.text in self.item_unit:
                prev.find_by_tag('input').fill(self.item_count)
                return True
            prev = td

        return False


def main():
    log.write(f"{datetime.now().strftime('%A %B %d, %Y %I:%M %p')}\n")
    log.write('Log start\n\n')

    wb = load_workbook(filename=FILE, read_only=True)
    sheet = wb.active

    # Visit portal and log in
    browser.visit(URL)
    browser.driver.maximize_window()
    browser.fill('username', USER)
    browser.fill('PASSWORD', PASS)
    browser.find_by_id('submit').click()

    # Wait for redirect
    while not browser.url.endswith('erslaunch-app'):
        time.sleep(1)

    # Navigate to inventory page
    tags = browser.find_by_tag('h3')
    find_and_click(tags, search_type='text', search_text='Enterprise Office')

    wait_for_load()

    browser.find_by_text('Shortcuts').click()
    tags = browser.find_by_css('.style3')
    find_and_click(tags, search_type='text', search_text='Inventory')

    wait_for_load()
    
    if NEW_INV:
        # Add new inventory sheet
        browser.find_by_id('ADD_ACTION').click()
        browser.find_by_css('.selectBox-arrow').last.click()
        links = browser.find_by_tag('li').links.find_by_text(FREQ)
        find_and_click(links, search_type='visible')
        browser.fill('DATE_1', DATE)
        browser.find_by_value('Add').click()
    else:
        # Find correct inventory sheet
        browser.find_by_css('.controls').click()
        links = browser.find_by_tag('li').links.find_by_text(FREQ)
        find_and_click(links, search_type='visible')
        browser.fill('DATE_2', DATE)
        browser.fill('DATE_3', DATE)
        browser.find_by_value('GO').click()

        wait_for_load()

        links = browser.find_by_name('openObject')
        find_and_click(links, search_type='visible')

    wait_for_load()

    # Cycle through spreadsheet and enter data
    for row in sheet.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=0, max_col=8):
        item = Item()
        if not item.parse_row(row): continue

        log_text = f"ADDED: {item}\n" if item.enter_data() else f"\nERROR: {item}  WAS NOT FOUND\n\n"
        log.write(log_text)
    
    if AUTO_SAVE:
        buttons = browser.find_by_value('Save')
        find_and_click(buttons, search_type='visible')

        wait_for_load()

        browser.find_by_text('OK').click()


if __name__ == '__main__':
    try:
        FREQ, DATE, FILE, NEW_INV = startup_gui()
    except TypeError:
        quit()

    log = open(f"{PATH}/log.txt", 'a')
    log.truncate(0)

    try:
        if LEGACY:
            CHROME_SERVICE = ChromeService(executable_path=f"{PATH}/chromedriver_win32/chromedriver")
            browser = Browser(BROWSER_NAME, service=CHROME_SERVICE)
        else:
            browser = Browser(BROWSER_NAME)

        main()
    except Exception as e:
        traceback.print_exc()
        log.write(f"{str(e)}\n")
        log.write(f"{traceback.format_exc()}")
    finally:
        log.write("\nLog closed")
        log.close()
        browser.quit()