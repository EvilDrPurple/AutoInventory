import configparser
import os
import re
import subprocess
import sys
import time
import traceback
from datetime import datetime

import PySimpleGUI as sg
import requests
from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from selenium.webdriver.chrome.service import Service as ChromeService
from splinter import Browser
from splinter.exceptions import ElementDoesNotExist

from exceptions import LoginFailedError, UserCancelled

config = configparser.ConfigParser()
config.read('config.ini')

VERSION = open('VERSION.txt', 'r').readline().strip()
URL = 'https://fedsso.yum.com/idp/startSSO.ping?PartnerSpId=https://yumph.altametrics.com/'
UNITS = {'EACH': {'DISK', 'EACH'},
        'BTL': 'BOTTLE',
        'GAL': 'GALLON'}
USER = config['Login Details']['username']
PASS = config['Login Details']['password']
MIN_ROW = config.getint('Spreadsheet', 'min_row')
MAX_ROW = config.getint('Spreadsheet', 'max_row')
AUTO_SAVE = config.getboolean('Important Things', 'auto_save')
LEGACY = config.getboolean('Important Things', 'legacy')
BROWSER_NAME = config['Important Things']['browser']


def update_program():
    """Checks for updates to the program, calls the updater executable if one is available."""
    
    VERSION_URL = 'https://raw.githubusercontent.com/EvilDrPurple/AutoInventory/master/VERSION.txt'
    REMOTE_VERSION = requests.get(VERSION_URL).text.strip()

    if (REMOTE_VERSION > VERSION):
        subprocess.Popen(f"update.exe {LEGACY}")
        sys.exit()


def settings_menu(first_time=False):
    global USER, PASS
    FONT = 'Ariel 14'
    settings_layout = [  [sg.Text('Enter your eResturant login details', font=FONT)],
            [sg.Text('Username:', font=FONT), sg.Push(), sg.Input(key='-USER-', default_text=USER, size=25)],
            [sg.Text('Password:', font=FONT), sg.Push(), sg.Input(key='-PASS-', default_text=PASS, size=25)],
            [sg.Text()],
            [sg.Push(), sg.Ok(font=FONT), sg.Cancel(font=FONT), sg.Push()] ]

    settings_window = sg.Window('Settings', settings_layout)

    while True:
        event, values = settings_window.read()

        if event in (sg.WIN_CLOSED, 'Cancel'):
            settings_window.close()
            if first_time: break
            else: return

        if first_time and (not values['-USER-'] or not values['-PASS-']):
            popup('An eResturant username and password is required', title='Oopsie')
        else:
            config['Login Details']['username'] = values['-USER-']
            config['Login Details']['password'] = values['-PASS-']

            with open('config.ini', 'w') as configfile:
                config.write(configfile)

            USER, PASS = values['-USER-'], values['-PASS-']
            settings_window.close()
            return

    sys.exit()


def startup_gui():
    """GUI that runs on program startup.

    Allows the user to specify values for inventory frequency, inventory date, filepath to inventory spreadsheet,
    and whether a new inventory sheet should be created.

    Returns:
        tuple(str, str, str, bool): A tuple that contains user-specified values of frequency, date, filepath, and new_inv.
    """

    sg.theme('DarkPurple4')
    sg.theme_text_color('white')

    if not USER or not PASS: settings_menu(True)

    FONT = 'Ariel 14'
    TOOLTIP = 'Uncheck this option if there is already an existing inventory sheet for the specified date and frequency in eResturant'
    OPTION_MENU = sg.OptionMenu(['Daily', 'Weekly', 'Monthly'], default_value='Weekly', key='-FREQ-')
    CALENDAR_BUTTON = sg.CalendarButton('Select Date', target='-DATE-', format='%m/%d/%Y')
    CHECKBOX = sg.Checkbox('Create new inventory sheet', font=FONT, default=True, key='-NEW_INV-', tooltip=TOOLTIP)
    IMAGE_SETTINGS = './Images/settings_button.png'
    SETTINGS_BUTTON = sg.Button(image_filename=IMAGE_SETTINGS, image_size=(33, 33), image_subsample=3, key='-SETTINGS-')

    layout = [  [sg.Text('Select count frequency:', font=FONT), sg.Push(), OPTION_MENU],
                [sg.Text('Enter date (mm/dd/yyyy):', font=FONT), sg.Push(), sg.Input(key='-DATE-', size=12), CALENDAR_BUTTON],
                [sg.Text('Select inventory spreadsheet:', font=FONT)],
                [sg.Input(key='-FILE-', size=50), sg.FileBrowse(file_types=(('Microsoft Excel Worksheet', '*.xlsx'),))],
                [sg.Push(), CHECKBOX, sg.Push()],
                [sg.Text()],
                [sg.Push(), sg.Ok(font=FONT), sg.Cancel(font=FONT), sg.Push(), SETTINGS_BUTTON] ]

    window = sg.Window(f"AutoInventory - v{VERSION}", layout)

    vali_date = lambda date : re.search('^(1[0-2]|0?[1-9])/(0?[1-9]|[1-2]\d|3[0-1])/\d{4}$', date)

    while True:
        event, values = window.read()

        if event in (sg.WIN_CLOSED, 'Cancel'):
            break
        
        if event == '-SETTINGS-':
            settings_menu()
        elif not vali_date(values['-DATE-']):
            popup('Please enter a valid date', title='Invalid date')
        elif not values['-FILE-'].endswith('.xlsx'):
            popup('Please select a valid file', title='Invalid file')
        else:
            window.close()
            return values['-FREQ-'], values['-DATE-'], values['-FILE-'], values['-NEW_INV-']

    window.close()


def popup(text, title, button_type=sg.POPUP_BUTTONS_OK):
    """Displays a popup window.

    Args:
        text (string): Message to display.
        title (string): Title of the popup window.
        button_type (int, optional): Button layout type. Defaults to sg.POPUP_BUTTONS_OK.

    Returns:
        str | None: Text of the button that was pressed or None if closed.
    """

    return sg.popup(text, title=title, button_type=button_type, font='Ariel 14', keep_on_top=True)


def login():
    """Visits and logs into the eResturant website.

    Raises:
        LoginFailedError: If the login fails.
    """

    browser.visit(URL)
    browser.driver.maximize_window()
    
    browser.fill('username', USER)
    browser.fill('PASSWORD', PASS)
    browser.find_by_id('submit').click()

    if 'pkmslogin' in browser.url and browser.find_by_id('errorMSG').visible:
        popup('Please update your login information in the config file', title='Login Failed')
        raise LoginFailedError(USER)


def wait_for_load():
    """Waits for the page to load."""    

    try:
        if browser.find_by_id('loading_layer').is_visible(0.5): wait_for_load()
    except (ElementDoesNotExist, IndexError):
        time.sleep(0.5)
        wait_for_load()

def find_and_click(items, search_type, search_text=''):
    """Finds and clicks on specified object

    This function was created because sometimes the find_by functions of splinter would be unable to find an element or finds an undesired element.
    This function corrects the problem by searching through an element's text or by only clicking on elements that are visible.

    Args:
        items (ElementList): List of elements to search through.
        search_type (str): Determines search behavior, either 'text' or 'visible'.
        search_text (str, optional): Text to search for, only valid when search_type is 'text'. Defaults to ''.

    Returns:
        bool: True if an element is found, False otherwise.
    """

    for item in items:
        if (
            search_type == 'text' and item.text == search_text
            or search_type == 'visible' and item.visible
        ):
            item.click()
            return True
        
    return False


def create_inventory_sheet():
    """Creates a new inventory sheet in eResturant.

    If an inventory sheet already exists for the specified date and frequency, a popup will appear asking if the user
    would like to use the existing inventory sheet. If yes, open_inventory_sheet() will be called.

    Raises:
        UserCancelled: If user cancels in popup window.
    """

    browser.find_by_id('ADD_ACTION').click()
    browser.find_by_css('.selectBox-arrow').last.click()
    links = browser.find_by_tag('li').links.find_by_text(FREQ)
    find_and_click(links, search_type='visible')
    browser.fill('DATE_1', DATE)
    browser.find_by_value('Add').click()

    wait_for_load()

    if browser.find_by_id('pop_msg').is_visible(0.5):
        result = popup(f"{FREQ} inventory sheet already exists for the date {DATE}\nWould you like to use it?",
                       title='Inventory sheet already exists', button_type=sg.POPUP_BUTTONS_YES_NO)
        
        if result == 'Yes':
            browser.find_by_text('OK').click()
            open_inventory_sheet()
        else: 
            raise UserCancelled()


def open_inventory_sheet():
    """Opens an inventory sheet in eResturant.

    If an inventory sheet does not exists for the specified date and frequency, a popup will appear asking if the user
    would like to create a new inventory sheet. If yes, create_inventory_sheet() will be called.

    Raises:
        UserCancelled: If user cancels in popup window.
    """

    browser.find_by_css('.controls').click()
    links = browser.find_by_tag('li').links.find_by_text(FREQ)
    find_and_click(links, search_type='visible')
    browser.fill('DATE_2', DATE)
    browser.fill('DATE_3', DATE)
    browser.find_by_value('GO').click()

    wait_for_load()

    links = browser.find_by_name('openObject')
    if not find_and_click(links, search_type='visible'):
        result = popup(f"No {FREQ} inventory sheet was found for the date {DATE}\nWould you like to create one?",
                       title='Unable to find inventory sheet', button_type=sg.POPUP_BUTTONS_YES_NO)
        
        if result == 'Yes':
            create_inventory_sheet()
        else: 
            raise UserCancelled()


def save_inventory_sheet():
    """Clicks on the save button"""

    buttons = browser.find_by_value('Save')
    find_and_click(buttons, search_type='visible')

    wait_for_load()

    browser.find_by_text('OK').click()


class Item:
    """Contains relevant information for inventory item.

    Attributes:
        item_code (str): Item code or id.
        item_desc (str): Item description or name.
        item_unit (str): Item unit such as LB.
        item_count (str): Item quantity.
    """

    def __init__(self):
        self.item_code = None

    def __str__(self):
        return f"{self.item_code : <12}{self.item_desc : <30}{self.item_count} {self.item_unit}"

    def parse_row(self, row):
        """Stores data from a given row of cells in the object's attributes.

        Args:
            row (list): List of Cell objects for the current row.

        Returns:
            bool: True if row is valid and all relevant date is accounted for, False otherwise.
        """

        for cell in row:
            if type(cell) is EmptyCell: continue
            match cell.column_letter:
                case 'A':
                    if not cell.value: return False
                    self.item_code = cell.value.strip()
                case 'B':
                    self.item_desc = cell.value.strip().replace('\t', '')
                    # Adjustment for Pepsi gallons
                    self.item_code = 'V62 Syrup' if self.item_desc == 'BNB PEPSI 5 GL SYRUP' else self.item_code
                    self.item_code = 'V65 Syrup' if self.item_desc == 'BNB PEPSI 3 GL SYRUP' else self.item_code
                case 'C':
                    if not cell.value: return False
                    self.item_unit = cell.value.strip()
                    # Adjustment for oregano
                    self.item_unit = 'CASE' if self.item_code == '74727' else self.item_unit
                    self.item_unit = UNITS[self.item_unit] if self.item_unit in UNITS else self.item_unit
                case 'H':
                    if not cell.value: return False
                    self.item_count = str(cell.value).strip()

        if not self.item_code: return False
        return True
    
    def enter_data(self):
        """Searches for item_code and enters the item_count into the proper field using item_unit.

        Returns:
            bool: True if the item_count is successfully entered, False otherwise.
        """

        browser.find_by_id('INV_ACC_DETAIL_tbl_filter').find_by_tag('input').fill(self.item_code)

        for td in browser.find_by_id('INV_ACC_DETAIL_tbl').find_by_tag('td'):
            if td.text and td.text in self.item_unit:
                prev.find_by_tag('input').fill(self.item_count)
                return True
            prev = td

        return False


def main():
    log.write(f"{datetime.now().strftime('%A %B %d, %Y %I:%M %p')}\n")
    log.write('Log start\n\n')

    wb = load_workbook(filename=FILE, read_only=True)
    sheet = wb.active

    # Visit portal and log in
    login()

    # Wait for redirect
    while not browser.url.endswith('erslaunch-app'): time.sleep(1)

    # Navigate to inventory page
    tags = browser.find_by_tag('h3')
    find_and_click(tags, search_type='text', search_text='Enterprise Office')

    wait_for_load()

    browser.find_by_text('Shortcuts').click()
    tags = browser.find_by_css('.style3')
    find_and_click(tags, search_type='text', search_text='Inventory')

    wait_for_load()
    
    # Create or open inventory sheet
    try: create_inventory_sheet() if NEW_INV else open_inventory_sheet()
    except (IndexError, ElementDoesNotExist): pass

    wait_for_load()

    # Cycle through spreadsheet and enter data
    for row in sheet.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=0, max_col=8):
        item = Item()
        if not item.parse_row(row): continue

        log_text = f"ADDED: {item}\n" if item.enter_data() else f"\nERROR: {item}  WAS NOT FOUND\n\n"
        log.write(log_text)
    
    if AUTO_SAVE: save_inventory_sheet()

    popup('Be sure to check the log file and any warnings in eResturant before posting', title='Saved successfully')


if __name__ == '__main__':
    update_program()

    try:
        FREQ, DATE, FILE, NEW_INV = startup_gui()
    except TypeError:
        sys.exit()

    log = open('log.txt', 'a')
    log.truncate(0)

    if LEGACY:
        CHROME_SERVICE = ChromeService()
        browser = Browser(BROWSER_NAME, service=CHROME_SERVICE)
    else:
        browser = Browser(BROWSER_NAME)

    try:
        main()
    except (LoginFailedError, UserCancelled) as e:
        log.write(e.message + '\n')
    except Exception as e:
        traceback.print_exc()
        log.write(str(e) + '\n')
        log.write(traceback.format_exc())
    finally:
        log.write('\nLog closed')
        log.close()
        browser.quit()
        os.system('notepad log.txt')
